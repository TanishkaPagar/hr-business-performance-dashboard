import random
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date, timedelta

random.seed(42)
np.random.seed(42)

# ── CONFIG ────────────────────────────────────────────────────────────────────
N = 3000
DEPARTMENTS   = ['Engineering', 'Sales', 'HR', 'Finance', 'Marketing', 'Operations', 'Legal', 'Product']
JOB_ROLES     = {
    'Engineering': ['Software Engineer', 'Senior Engineer', 'Tech Lead', 'QA Engineer'],
    'Sales':       ['Sales Executive', 'Account Manager', 'Sales Manager', 'BDR'],
    'HR':          ['HR Generalist', 'Recruiter', 'HR Manager', 'L&D Specialist'],
    'Finance':     ['Financial Analyst', 'Accountant', 'Finance Manager', 'Controller'],
    'Marketing':   ['Marketing Analyst', 'Content Strategist', 'SEO Specialist', 'Brand Manager'],
    'Operations':  ['Operations Analyst', 'Supply Chain Lead', 'Process Manager', 'Logistics Coordinator'],
    'Legal':       ['Legal Counsel', 'Compliance Officer', 'Paralegal', 'Legal Manager'],
    'Product':     ['Product Manager', 'Product Analyst', 'UX Designer', 'Scrum Master'],
}
LOCATIONS     = ['Mumbai', 'Delhi', 'Bangalore', 'Hyderabad', 'Pune', 'Chennai', 'Kolkata']
BUSINESS_UNITS= ['North India', 'South India', 'West India', 'East India', 'Central']
SALARY_BANDS  = ['Band 1 (Entry)', 'Band 2 (Mid)', 'Band 3 (Senior)', 'Band 4 (Lead)', 'Band 5 (Executive)']
EDUCATION     = ['High School', 'Bachelor\'s', 'Master\'s', 'MBA', 'PhD']

def salary_for_band(band):
    ranges = {
        'Band 1 (Entry)':      (25000,  45000),
        'Band 2 (Mid)':        (45000,  75000),
        'Band 3 (Senior)':     (75000, 120000),
        'Band 4 (Lead)':       (120000,180000),
        'Band 5 (Executive)':  (180000,300000),
    }
    lo, hi = ranges[band]
    return round(random.uniform(lo, hi), -2)

def band_for_tenure(tenure):
    if tenure < 2:   return 'Band 1 (Entry)'
    if tenure < 5:   return random.choice(['Band 1 (Entry)', 'Band 2 (Mid)'])
    if tenure < 9:   return random.choice(['Band 2 (Mid)', 'Band 3 (Senior)'])
    if tenure < 14:  return random.choice(['Band 3 (Senior)', 'Band 4 (Lead)'])
    return random.choice(['Band 4 (Lead)', 'Band 5 (Executive)'])

# ── GENERATE RAW DATA ─────────────────────────────────────────────────────────
rows = []
for i in range(1, N + 1):
    dept   = random.choice(DEPARTMENTS)
    role   = random.choice(JOB_ROLES[dept])
    tenure = round(random.uniform(0.5, 20), 1)
    band   = band_for_tenure(tenure)
    salary = salary_for_band(band)
    perf   = round(random.gauss(3.2, 0.7), 1)
    perf   = max(1.0, min(5.0, perf))
    engage = round(random.gauss(65, 15), 1)
    engage = max(10.0, min(100.0, engage))
    wlb    = random.randint(1, 5)
    train  = random.randint(5, 120)
    target = round(random.gauss(85, 15), 1)
    target = max(20.0, min(130.0, target))
    projects = random.randint(1, 20)
    bonus  = round(random.uniform(0, 0.25) * (perf / 5), 4)
    hire_date = date.today() - timedelta(days=int(tenure * 365))
    last_promo = hire_date + timedelta(days=random.randint(180, max(181, int(tenure * 300))))
    last_promo = min(last_promo, date.today())
    # attrition: higher for low engagement, low perf, entry band
    attrition_prob = 0.05 + (0.15 if engage < 50 else 0) + (0.10 if perf < 2.5 else 0) + \
                     (0.08 if band == 'Band 1 (Entry)' else 0)
    attrition = 'Yes' if random.random() < attrition_prob else 'No'
    gender = random.choice(['Male', 'Female', 'Non-Binary'])
    age    = max(22, min(60, int(22 + tenure + random.uniform(0, 15))))
    edu    = random.choices(EDUCATION, weights=[5, 40, 30, 20, 5])[0]
    rev_contrib = round(salary * random.uniform(1.5, 6.0), -3)

    rows.append({
        'EmployeeID':          f'EMP{i:04d}',
        'Gender':              gender,
        'Age':                 age,
        'Education':           edu,
        'Department':          dept,
        'JobRole':             role,
        'Location':            random.choice(LOCATIONS),
        'BusinessUnit':        random.choice(BUSINESS_UNITS),
        'HireDate':            hire_date.strftime('%Y-%m-%d'),
        'TenureYears':         tenure,
        'SalaryBand':          band,
        'MonthlySalary':       salary,
        'BonusPercent':        round(bonus * 100, 2),
        'PerformanceRating':   perf,
        'TargetAchievementPct':target,
        'TrainingHours':       train,
        'EngagementScore':     engage,
        'JobSatisfaction':     random.randint(1, 5),
        'WorkLifeBalance':     wlb,
        'LastPromotionDate':   last_promo.strftime('%Y-%m-%d'),
        'ProjectsCompleted':   projects,
        'RevenueContribution': rev_contrib,
        'Attrition':           attrition,
    })

df = pd.DataFrame(rows)

# ── SUMMARY TABLES ────────────────────────────────────────────────────────────
def attrition_rate(d):
    return round(d['Attrition'].eq('Yes').mean() * 100, 2)

dept_summary = df.groupby('Department').agg(
    Headcount       = ('EmployeeID', 'count'),
    AttritionCount  = ('Attrition', lambda x: (x == 'Yes').sum()),
    AvgSalary       = ('MonthlySalary', 'mean'),
    AvgPerformance  = ('PerformanceRating', 'mean'),
    AvgEngagement   = ('EngagementScore', 'mean'),
    AvgTrainingHrs  = ('TrainingHours', 'mean'),
    TotalRevenue    = ('RevenueContribution', 'sum'),
).reset_index()
dept_summary['AttritionRate%'] = (dept_summary['AttritionCount'] / dept_summary['Headcount'] * 100).round(2)
dept_summary['AvgSalary']      = dept_summary['AvgSalary'].round(0)
dept_summary['AvgPerformance'] = dept_summary['AvgPerformance'].round(2)
dept_summary['AvgEngagement']  = dept_summary['AvgEngagement'].round(2)
dept_summary['AvgTrainingHrs'] = dept_summary['AvgTrainingHrs'].round(1)

salary_band_summary = df.groupby('SalaryBand').agg(
    Headcount      = ('EmployeeID', 'count'),
    AvgSalary      = ('MonthlySalary', 'mean'),
    MinSalary      = ('MonthlySalary', 'min'),
    MaxSalary      = ('MonthlySalary', 'max'),
    AttritionCount = ('Attrition', lambda x: (x == 'Yes').sum()),
).reset_index()
salary_band_summary['AttritionRate%'] = (salary_band_summary['AttritionCount'] / salary_band_summary['Headcount'] * 100).round(2)
salary_band_summary['AvgSalary'] = salary_band_summary['AvgSalary'].round(0)
salary_band_summary['MinSalary'] = salary_band_summary['MinSalary'].round(0)
salary_band_summary['MaxSalary'] = salary_band_summary['MaxSalary'].round(0)

age_bins   = [21, 30, 40, 50, 61]
age_labels = ['22-30', '31-40', '41-50', '51-60']
df['AgeGroup'] = pd.cut(df['Age'], bins=age_bins, labels=age_labels, right=True)
age_summary = df.groupby('AgeGroup', observed=True).agg(
    Headcount      = ('EmployeeID', 'count'),
    AttritionCount = ('Attrition', lambda x: (x == 'Yes').sum()),
    AvgSalary      = ('MonthlySalary', 'mean'),
    AvgPerformance = ('PerformanceRating', 'mean'),
).reset_index()
age_summary['AttritionRate%'] = (age_summary['AttritionCount'] / age_summary['Headcount'] * 100).round(2)
age_summary['AvgSalary']      = age_summary['AvgSalary'].round(0)
age_summary['AvgPerformance'] = age_summary['AvgPerformance'].round(2)
age_summary['AgeGroup']       = age_summary['AgeGroup'].astype(str)

bu_summary = df.groupby('BusinessUnit').agg(
    Headcount       = ('EmployeeID', 'count'),
    AttritionCount  = ('Attrition', lambda x: (x == 'Yes').sum()),
    AvgPerformance  = ('PerformanceRating', 'mean'),
    TotalRevenue    = ('RevenueContribution', 'sum'),
    ProjectsTotal   = ('ProjectsCompleted', 'sum'),
).reset_index()
bu_summary['AttritionRate%'] = (bu_summary['AttritionCount'] / bu_summary['Headcount'] * 100).round(2)
bu_summary['AvgPerformance'] = bu_summary['AvgPerformance'].round(2)

# ── WRITE EXCEL ───────────────────────────────────────────────────────────────
wb = Workbook()

# Styles
HEADER_FONT    = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL_1  = PatternFill('solid', start_color='1F4E79')  # dark blue
HEADER_FILL_2  = PatternFill('solid', start_color='2E75B6')  # medium blue
HEADER_FILL_3  = PatternFill('solid', start_color='ED7D31')  # orange
HEADER_FILL_4  = PatternFill('solid', start_color='70AD47')  # green
HEADER_FILL_5  = PatternFill('solid', start_color='7030A0')  # purple
ALT_FILL       = PatternFill('solid', start_color='D9E1F2')
CENTER         = Alignment(horizontal='center', vertical='center')
LEFT           = Alignment(horizontal='left',   vertical='center')
THIN           = Side(style='thin', color='BFBFBF')
BORDER         = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header_row(ws, row_num, fill, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font      = HEADER_FONT
        cell.fill      = fill
        cell.alignment = CENTER
        cell.border    = BORDER

def style_data_rows(ws, start_row, end_row, col_count):
    for r in range(start_row, end_row + 1):
        fill = ALT_FILL if r % 2 == 0 else PatternFill()
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill      = fill
            cell.alignment = LEFT
            cell.border    = BORDER
            cell.font      = Font(name='Arial', size=10)

def write_df_to_sheet(ws, df_in, fill, start_row=1):
    cols = list(df_in.columns)
    for ci, col in enumerate(cols, 1):
        ws.cell(row=start_row, column=ci, value=col)
    style_header_row(ws, start_row, fill, len(cols))
    for ri, row in enumerate(df_in.itertuples(index=False), start_row + 1):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    style_data_rows(ws, start_row + 1, start_row + len(df_in), len(cols))
    for ci, col in enumerate(cols, 1):
        max_len = max(len(str(col)), df_in[col].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 30)
    return start_row + len(df_in) + 1

# ── SHEET 1: Raw Data ─────────────────────────────────────────────────────────
ws_raw = wb.active
ws_raw.title = 'Raw Data'
raw_export = df.drop(columns=['AgeGroup'])
write_df_to_sheet(ws_raw, raw_export, HEADER_FILL_1)
ws_raw.freeze_panes = 'A2'

# ── SHEET 2: KPI Summary ──────────────────────────────────────────────────────
ws_kpi = wb.create_sheet('KPI Summary')
ws_kpi['A1'] = 'HR & Business Performance — KPI Summary'
ws_kpi['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
ws_kpi['A1'].alignment = CENTER
ws_kpi.merge_cells('A1:D1')

kpis = [
    ('Total Employees',       len(df)),
    ('Attrition Rate %',      f"=COUNTIF('Raw Data'!W2:W{N+1},\"Yes\")/COUNTA('Raw Data'!A2:A{N+1})*100"),
    ('Avg Monthly Salary',    f"=AVERAGE('Raw Data'!L2:L{N+1})"),
    ('Avg Performance Rating',f"=AVERAGE('Raw Data'!N2:N{N+1})"),
    ('Avg Engagement Score',  f"=AVERAGE('Raw Data'!Q2:Q{N+1})"),
    ('Avg Training Hours',    f"=AVERAGE('Raw Data'!P2:P{N+1})"),
    ('Total Revenue (₹)',     f"=SUM('Raw Data'!V2:V{N+1})"),
    ('Total Projects Completed', f"=SUM('Raw Data'!U2:U{N+1})"),
]
ws_kpi['A3'] = 'KPI'; ws_kpi['B3'] = 'Value'
style_header_row(ws_kpi, 3, HEADER_FILL_1, 2)
for i, (kpi, val) in enumerate(kpis, 4):
    ws_kpi.cell(row=i, column=1, value=kpi).font  = Font(name='Arial', bold=True, size=11)
    ws_kpi.cell(row=i, column=2, value=val).font   = Font(name='Arial', size=11)
    ws_kpi.cell(row=i, column=1).border = BORDER
    ws_kpi.cell(row=i, column=2).border = BORDER
    if i % 2 == 0:
        ws_kpi.cell(row=i, column=1).fill = ALT_FILL
        ws_kpi.cell(row=i, column=2).fill = ALT_FILL
ws_kpi.column_dimensions['A'].width = 30
ws_kpi.column_dimensions['B'].width = 22

# ── SHEET 3: Department Summary ───────────────────────────────────────────────
ws_dept = wb.create_sheet('Dept Summary')
ws_dept['A1'] = 'Department-wise HR & Performance Summary'
ws_dept['A1'].font = Font(name='Arial', bold=True, size=13, color='1F4E79')
ws_dept.merge_cells(f'A1:{get_column_letter(len(dept_summary.columns))}1')
write_df_to_sheet(ws_dept, dept_summary, HEADER_FILL_2, start_row=3)

# ── SHEET 4: Salary Band Summary ─────────────────────────────────────────────
ws_sal = wb.create_sheet('Salary Band Summary')
ws_sal['A1'] = 'Salary Band Analysis'
ws_sal['A1'].font = Font(name='Arial', bold=True, size=13, color='1F4E79')
ws_sal.merge_cells(f'A1:{get_column_letter(len(salary_band_summary.columns))}1')
write_df_to_sheet(ws_sal, salary_band_summary, HEADER_FILL_3, start_row=3)

# ── SHEET 5: Age Group Summary ────────────────────────────────────────────────
ws_age = wb.create_sheet('Age Group Summary')
ws_age['A1'] = 'Attrition & Performance by Age Group'
ws_age['A1'].font = Font(name='Arial', bold=True, size=13, color='1F4E79')
ws_age.merge_cells(f'A1:{get_column_letter(len(age_summary.columns))}1')
write_df_to_sheet(ws_age, age_summary, HEADER_FILL_4, start_row=3)

# ── SHEET 6: Business Unit Summary ───────────────────────────────────────────
ws_bu = wb.create_sheet('Business Unit Summary')
ws_bu['A1'] = 'Business Unit Performance Overview'
ws_bu['A1'].font = Font(name='Arial', bold=True, size=13, color='1F4E79')
ws_bu.merge_cells(f'A1:{get_column_letter(len(bu_summary.columns))}1')
write_df_to_sheet(ws_bu, bu_summary, HEADER_FILL_5, start_row=3)

# ── SHEET 7: Data Dictionary ──────────────────────────────────────────────────
ws_dict = wb.create_sheet('Data Dictionary')
ws_dict['A1'] = 'Data Dictionary — HR & Business Performance Dataset'
ws_dict['A1'].font = Font(name='Arial', bold=True, size=13, color='1F4E79')
ws_dict.merge_cells('A1:C1')
dict_data = [
    ('Column', 'Type', 'Description'),
    ('EmployeeID',           'Text',    'Unique employee identifier (EMP0001–EMP3000)'),
    ('Gender',               'Text',    'Male / Female / Non-Binary'),
    ('Age',                  'Integer', 'Employee age in years (22–60)'),
    ('Education',            'Text',    'Highest education level attained'),
    ('Department',           'Text',    'One of 8 departments'),
    ('JobRole',              'Text',    'Specific role within department'),
    ('Location',             'Text',    'Office city (7 Indian cities)'),
    ('BusinessUnit',         'Text',    'Regional business unit (5 regions)'),
    ('HireDate',             'Date',    'Date of joining (YYYY-MM-DD)'),
    ('TenureYears',          'Decimal', 'Years with the company'),
    ('SalaryBand',           'Text',    'Band 1 (Entry) to Band 5 (Executive)'),
    ('MonthlySalary',        'Number',  'Gross monthly salary in INR'),
    ('BonusPercent',         'Decimal', 'Annual bonus as % of salary (0–25%)'),
    ('PerformanceRating',    'Decimal', 'Annual rating 1.0–5.0 (5 = Outstanding)'),
    ('TargetAchievementPct', 'Decimal', 'Business target achieved % (20–130%)'),
    ('TrainingHours',        'Integer', 'Training hours completed in the year'),
    ('EngagementScore',      'Decimal', 'Employee engagement score 10–100'),
    ('JobSatisfaction',      'Integer', 'Self-reported satisfaction 1–5'),
    ('WorkLifeBalance',      'Integer', 'Self-reported WLB score 1–5'),
    ('LastPromotionDate',    'Date',    'Date of last promotion (YYYY-MM-DD)'),
    ('ProjectsCompleted',    'Integer', 'Projects completed in the year (1–20)'),
    ('RevenueContribution',  'Number',  'Estimated revenue contribution in INR'),
    ('Attrition',            'Text',    'Yes = left the company, No = still active'),
]
for ri, row in enumerate(dict_data, 3):
    for ci, val in enumerate(row, 1):
        cell = ws_dict.cell(row=ri, column=ci, value=val)
        cell.border = BORDER
        cell.font   = Font(name='Arial', bold=(ri == 3), size=10,
                           color='FFFFFF' if ri == 3 else '000000')
        if ri == 3:
            cell.fill = HEADER_FILL_1
        elif ri % 2 == 0:
            cell.fill = ALT_FILL
        cell.alignment = LEFT
ws_dict.column_dimensions['A'].width = 26
ws_dict.column_dimensions['B'].width = 12
ws_dict.column_dimensions['C'].width = 55

output_path = '/mnt/user-data/outputs/HR_Business_Performance_Dashboard.xlsx'
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Rows: {N} | Sheets: {len(wb.sheetnames)}")
